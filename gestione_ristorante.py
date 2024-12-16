import tkinter as tk
from tkinter import messagebox
import sqlite3
import openpyxl
import os
from openpyxl import Workbook

# Costanti per i giorni della settimana
GIORNI_SETTIMANA = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]

# Classe Dipendente per gestire le informazioni sui dipendenti
class Dipendente:
    def __init__(self, nome, ore_disponibili, ruoli=None):
        self.nome = nome
        self.ore_disponibili = ore_disponibili
        self.ruoli = ruoli if ruoli else []  # Lista di ruoli associati al dipendente

    def __str__(self):
        return f"{self.nome} - Ore disponibili: {self.ore_disponibili} - Ruoli: {', '.join(self.ruoli)}"

# Classe Ruolo per gestire le informazioni sui ruoli
class Ruolo:
    def __init__(self, nome):
        self.nome = nome

    def __str__(self):
        return f"Ruolo: {self.nome}"

# Classe Ristorante per gestire gli orari settimanali del ristorante
class Ristorante:
    def __init__(self, connection):
        self.connection = connection
        self.orari_settimanali = {}
        self.carica_orari_da_db()
    
    def carica_orari_da_db(self):
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT giorno, ore_lavorative FROM orari_settimanali")
            rows = cursor.fetchall()
            for row in rows:
                giorno, ore_lavorative = row
                self.orari_settimanali[giorno] = {"ore_lavorative": ore_lavorative}
        except sqlite3.Error as e:
            print("Errore nel caricamento degli orari")

    def modifica_orario_giorno(self, giorno, ore_lavorative):
        if giorno in GIORNI_SETTIMANA:
            self.orari_settimanali[giorno] = {"ore_lavorative": ore_lavorative}
            try:
                cursor = self.connection.cursor()
                cursor.execute(
                    "UPDATE orari_settimanali SET ore_lavorative=? WHERE giorno=?",
                    (ore_lavorative, giorno)
                )
                self.connection.commit()
                return True
            except sqlite3.Error as e:
                print("Errore nella modifica degli orari")
                return False
        return False

    def __str__(self):
        orari_str = "Orari Settimanali del Ristorante:\n"
        for giorno in GIORNI_SETTIMANA:
            orari = self.orari_settimanali.get(giorno, {"ore_lavorative": 0})
            orari_str += f"{giorno}: Ore lavorative: {orari['ore_lavorative']}\n"
        return orari_str

# Classe GestioneRistorante per gestire dipendenti, ruoli e stanze
class GestioneRistorante:
    def __init__(self, connection):
        self.connection = connection
        self.dipendenti = []
        self.stanze = []  # Lista di stanze
        self.ruoli = []  # Lista di ruoli
        self.num_persone_stanze_ruolo = {}  # Dizionario che associa stanze a numero di persone per ruolo
        self.orari = Ristorante(connection)
        self.carica_dipendenti_da_db()
        self.carica_ruoli_da_db()
        self.carica_stanze_da_db()

    # Carica dipendenti dal database
    def carica_dipendenti_da_db(self):
        try:
            cursor = self.connection.cursor()
            cursor.execute("""
                SELECT d.nome, d.ore_disponibili, r.nome 
                FROM dipendenti d
                LEFT JOIN dipendenti_ruoli dr ON d.id = dr.dipendente_id
                LEFT JOIN ruoli r ON dr.ruolo_id = r.id
            """)
            rows = cursor.fetchall()
            dipendenti_dict = {}
            for nome, ore_disponibili, ruolo in rows:
                if nome not in dipendenti_dict:
                    dipendenti_dict[nome] = Dipendente(nome, ore_disponibili, [])
                if ruolo:
                    dipendenti_dict[nome].ruoli.append(ruolo)
            self.dipendenti = list(dipendenti_dict.values())
        except sqlite3.Error as e:
            print("Errore nel caricamento dei dipendenti")

    def carica_ruoli_da_db(self):
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT nome FROM ruoli")
            rows = cursor.fetchall()
            self.ruoli = [Ruolo(nome) for nome, in rows]
        except sqlite3.Error as e:
            print("Errore nel caricamento dei ruoli")

    def __del__(self):
        self.connection.close()

    def carica_stanze_da_db(self):
        try:
            cursor = self.connection.cursor()

            # Carica tutte le stanze
            cursor.execute("SELECT id, nome FROM stanze")
            stanze_rows = cursor.fetchall()
            self.stanze = {row[0]: row[1] for row in stanze_rows}  # {id: nome}

            # Carica i ruoli associati alle stanze
            cursor.execute("""
                SELECT sr.stanza_id, r.nome, sr.numero_persone
                FROM stanze_ruoli sr
                JOIN ruoli r ON sr.ruolo_id = r.id
            """)
            stanze_ruoli_rows = cursor.fetchall()
            self.num_persone_stanze_ruolo = {}

            for stanza_id, ruolo_nome, numero_persone in stanze_ruoli_rows:
                stanza_nome = self.stanze[stanza_id]
                if stanza_nome not in self.num_persone_stanze_ruolo:
                    self.num_persone_stanze_ruolo[stanza_nome] = {}
                self.num_persone_stanze_ruolo[stanza_nome][ruolo_nome] = numero_persone

        except sqlite3.Error as e:
            print("Errore nel caricamento delle stanze")

    # CRUD Dipendenti
    def aggiungi_dipendente(self, nome, ore_disponibili, ruoli):
        try:
            cursor = self.connection.cursor()

            # Aggiungi il dipendente nella tabella dipendenti
            cursor.execute(
                "INSERT INTO dipendenti (nome, ore_disponibili) VALUES (?, ?)",
                (nome, ore_disponibili)
            )
            dipendente_id = cursor.lastrowid

            # Aggiungi i ruoli del dipendente nella tabella dipendenti_ruoli
            for ruolo_nome in ruoli:
                cursor.execute("SELECT id FROM ruoli WHERE nome = ?", (ruolo_nome,))
                ruolo_row = cursor.fetchone()
                if ruolo_row:
                    ruolo_id = ruolo_row[0]
                    cursor.execute(
                        "INSERT INTO dipendenti_ruoli (dipendente_id, ruolo_id) VALUES (?, ?)",
                        (dipendente_id, ruolo_id)
                    )
                else:
                    print(f"Ruolo '{ruolo_nome}' non trovato nel database")

            self.connection.commit()

            # Aggiorna la struttura in memoria
            dipendente = Dipendente(nome, ore_disponibili, ruoli)
            self.dipendenti.append(dipendente)

        except sqlite3.Error as e:
            print("Errore durante l'aggiunta del dipendente")
            self.connection.rollback()


    def rimuovi_dipendente(self, nome):
        try:
            cursor = self.connection.cursor()

            # Recupera l'ID del dipendente da eliminare
            cursor.execute("SELECT id FROM dipendenti WHERE nome=?", (nome,))
            dipendente_row = cursor.fetchone()

            if not dipendente_row:
                print(f"Dipendente '{nome}' non trovato nel database.")
                return False

            dipendente_id = dipendente_row[0]

            # Elimina il dipendente e i suoi ruoli associati
            cursor.execute("DELETE FROM dipendenti WHERE id=?", (dipendente_id,))
            cursor.execute("DELETE FROM dipendenti_ruoli WHERE dipendente_id=?", (dipendente_id,))
            self.connection.commit()

            # Aggiorna la struttura in memoria
            self.dipendenti = [d for d in self.dipendenti if d.nome != nome]
            return True

        except sqlite3.Error as e:
            print("Errore durante la rimozione del dipendente")
            self.connection.rollback()
            return False


    def modifica_dipendente(self, vecchio_nome, nuovo_nome, nuove_ore, nuovi_ruoli):
        try:
            cursor = self.connection.cursor()

            # Recupera l'ID del dipendente da modificare
            cursor.execute("SELECT id FROM dipendenti WHERE nome=?", (vecchio_nome,))
            dipendente_row = cursor.fetchone()

            if not dipendente_row:
                print(f"Dipendente '{vecchio_nome}' non trovato nel database.")
                return False

            dipendente_id = dipendente_row[0]

            # Aggiorna i dati del dipendente
            cursor.execute(
                "UPDATE dipendenti SET nome=?, ore_disponibili=? WHERE id=?",
                (nuovo_nome, nuove_ore, dipendente_id)
            )

            # Rimuovi i vecchi ruoli del dipendente
            cursor.execute("DELETE FROM dipendenti_ruoli WHERE dipendente_id=?", (dipendente_id,))

            # Aggiungi i nuovi ruoli del dipendente
            for ruolo_nome in nuovi_ruoli:
                cursor.execute("SELECT id FROM ruoli WHERE nome=?", (ruolo_nome,))
                ruolo_row = cursor.fetchone()
                if ruolo_row:
                    ruolo_id = ruolo_row[0]
                    cursor.execute(
                        "INSERT INTO dipendenti_ruoli (dipendente_id, ruolo_id) VALUES (?, ?)",
                        (dipendente_id, ruolo_id)
                    )
                else:
                    print(f"Ruolo '{ruolo_nome}' non trovato nel database.")

            self.connection.commit()

            # Aggiorna la struttura in memoria
            for dipendente in self.dipendenti:
                if dipendente.nome == vecchio_nome:
                    dipendente.nome = nuovo_nome
                    dipendente.ore_disponibili = nuove_ore
                    dipendente.ruoli = nuovi_ruoli
                    break

            return True

        except sqlite3.Error as e:
            print("Errore durante la modifica del dipendente")
            self.connection.rollback()
            return False


    # CRUD Ruoli
    def aggiungi_ruolo(self, nome):
        try:
            cursor = self.connection.cursor()
            cursor.execute("INSERT INTO ruoli (nome) VALUES (?)", (nome,))
            self.connection.commit()
            ruolo = Ruolo(nome)
            self.ruoli.append(ruolo)
        except sqlite3.Error as e:
            messagebox.showerror("Errore", f"Errore durante l'aggiunta del ruolo")
            self.connection.rollback()

    def rimuovi_ruolo(self, nome):
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT id FROM ruoli WHERE nome=?", (nome,))
            ruolo_row = cursor.fetchone()

            if not ruolo_row:
                print(f"Ruolo '{nome}' non trovato nel database.")
                return False

            ruolo_id = ruolo_row[0]
            cursor.execute("DELETE FROM dipendenti_ruoli WHERE ruolo_id=?", (ruolo_id,))
            cursor.execute("DELETE FROM stanze_ruoli WHERE ruolo_id=?", (ruolo_id,))
            cursor.execute("DELETE FROM ruoli WHERE id=?", (ruolo_id,))
            self.connection.commit()
            self.ruoli = [r for r in self.ruoli if r.nome != nome]
        except sqlite3.Error as e:
            print("Errore durante la rimozione del ruolo")
            self.connection.rollback()

    # CRUD Stanze
    def aggiungi_stanza(self, nome_stanza, persone_ruoli):
        try:
            cursor = self.connection.cursor()
            cursor.execute("INSERT INTO stanze (nome) VALUES (?)", (nome_stanza,))
            stanza_id = cursor.lastrowid

            for ruolo_nome, numero_persone in persone_ruoli.items():
                cursor.execute("SELECT id FROM ruoli WHERE nome=?", (ruolo_nome,))
                ruolo_row = cursor.fetchone()
                if ruolo_row:
                    ruolo_id = ruolo_row[0]
                    cursor.execute(
                        "INSERT INTO stanze_ruoli (stanza_id, ruolo_id, numero_persone) VALUES (?, ?, ?)",
                        (stanza_id, ruolo_id, numero_persone)
                    )

            self.connection.commit()
            self.stanze[stanza_id] = nome_stanza
            self.num_persone_stanze_ruolo[nome_stanza] = persone_ruoli
        except sqlite3.Error as e:
            print("Errore durante l'aggiunta della stanza")
            self.connection.rollback()

    def modifica_stanza(self, nome_stanza, persone_ruoli):
        try:
            cursor = self.connection.cursor()

            # Recupera l'ID della stanza
            stanza_id = next((id for id, nome in self.stanze.items() if nome == nome_stanza), None)

            if stanza_id is None:
                print(f"Stanza '{nome_stanza}' non trovata nel database.")
                return False

            # Rimuove i ruoli attuali associati alla stanza
            cursor.execute("DELETE FROM stanze_ruoli WHERE stanza_id=?", (stanza_id,))

            # Aggiunge i nuovi ruoli
            for ruolo_nome, numero_persone in persone_ruoli.items():
                cursor.execute("SELECT id FROM ruoli WHERE nome=?", (ruolo_nome,))
                ruolo_row = cursor.fetchone()
                if ruolo_row:
                    ruolo_id = ruolo_row[0]
                    cursor.execute(
                        "INSERT INTO stanze_ruoli (stanza_id, ruolo_id, numero_persone) VALUES (?, ?, ?)",
                        (stanza_id, ruolo_id, numero_persone)
                    )

            self.connection.commit()

            # Aggiorna la struttura in memoria
            self.num_persone_stanze_ruolo[nome_stanza] = persone_ruoli
            return True
        except sqlite3.Error as e:
            print("Errore durante la modifica della stanza:", e)
            self.connection.rollback()
            return False

    def rimuovi_stanza(self, nome_stanza):
        try:
            cursor = self.connection.cursor()
            stanza_id = next((id for id, nome in self.stanze.items() if nome == nome_stanza), None)

            if stanza_id is None:
                print(f"Stanza '{nome_stanza}' non trovata nel database.")
                return False

            cursor.execute("DELETE FROM stanze WHERE id=?", (stanza_id,))
            cursor.execute("DELETE FROM stanze_ruoli WHERE stanza_id=?", (stanza_id,))
            self.connection.commit()
            del self.stanze[stanza_id]
            del self.num_persone_stanze_ruolo[nome_stanza]
        except sqlite3.Error as e:
            print("Errore durante la rimozione della stanza")
            self.connection.rollback()

    # Pianificazione settimanale
    def genera_pianificazione_settimanale(self):
        pianificazione = []
        dipendenti_copia = [Dipendente(d.nome, d.ore_disponibili, d.ruoli) for d in self.dipendenti]

        for giorno in GIORNI_SETTIMANA:
            ore_giornaliere = self.calcola_ore_giornaliere(giorno)
            ore_giornaliere_rimanenti = ore_giornaliere
            giorno_pianificazione = []

            for stanza, persone_ruoli in self.num_persone_stanze_ruolo.items():
                for ruolo, persone_necessarie in persone_ruoli.items():
                    persone_assegnate = []
                    while persone_necessarie > 0 and ore_giornaliere_rimanenti > 0 and dipendenti_copia:
                        dipendente = next((d for d in dipendenti_copia if ruolo in d.ruoli), None)
                        if not dipendente:
                            break
                        ore_assegnabili = min(dipendente.ore_disponibili, ore_giornaliere_rimanenti)
                        if ore_assegnabili > 0:
                            persone_assegnate.append((dipendente.nome, ore_assegnabili, ruolo))
                            dipendente.ore_disponibili -= ore_assegnabili
                            ore_giornaliere_rimanenti -= ore_assegnabili
                            persone_necessarie -= 1

                    giorno_pianificazione.append((stanza, persone_assegnate))

            pianificazione.append((giorno, giorno_pianificazione))
        return pianificazione

    def calcola_ore_giornaliere(self, giorno):
        orari = self.orari.orari_settimanali[giorno]
        apertura_ore, apertura_minuti = map(int, orari['apertura'].split(':'))
        chiusura_ore, chiusura_minuti = map(int, orari['chiusura'].split(':'))
        apertura_totale_minuti = apertura_ore * 60 + apertura_minuti
        chiusura_totale_minuti = chiusura_ore * 60 + chiusura_minuti
        ore_giornaliere = (chiusura_totale_minuti - apertura_totale_minuti) / 60
        return max(0, ore_giornaliere)  # Assicura che le ore non siano negative

    def esporta_pianificazione_excel(self):
        try:
            pianificazione = []  # Pianificazione settimanale
            wb = Workbook()

            for giorno in GIORNI_SETTIMANA:
                ws = wb.create_sheet(title=giorno)
                ws.append(["Stanza", "Ruolo", "Numero Persone"])

                for stanza, persone_ruoli in self.num_persone_stanze_ruolo.items():
                    for ruolo, numero_persone in persone_ruoli.items():
                        ws.append([stanza, ruolo, numero_persone])

            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            download_path = os.path.join(os.path.expanduser("~"), "Downloads")
            file_path = os.path.join(download_path, "pianificazione_settimanale.xlsx")
            wb.save(file_path)
            messagebox.showinfo("Esportazione completata", f"La pianificazione è stata esportata in: {file_path}")
        except Exception as e:
            print("Errore durante l'esportazione")

class GestioneRistoranteApp:
    def __init__(self, root):
        # Inizializzazione della connessione al database
        self.connection = sqlite3.connect('gestione_ristorante.db')
        self.crea_tabelle()
        self.gestione = GestioneRistorante(self.connection)

        # Configurazione della finestra principale
        self.root = root
        self.root.title("Gestione Ristorante")

        # Configura la dimensione della finestra per adattarsi allo schermo
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        window_width = int(screen_width * 0.75)
        window_height = int(screen_height * 0.75)
        root.geometry(f"{window_width}x{window_height}")

        # Configurazione del layout a griglia per distribuire le colonne uniformemente
        self.root.grid_columnconfigure(0, weight=1, uniform="column")
        self.root.grid_columnconfigure(1, weight=1, uniform="column")
        self.root.grid_columnconfigure(2, weight=1, uniform="column")

        # Creazione dei frame per ciascuna sezione
        self.frame_personale = tk.Frame(self.root, bd=2, relief="groove")
        self.frame_personale.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)

        self.frame_stanze = tk.Frame(self.root, bd=2, relief="groove")
        self.frame_stanze.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)

        self.frame_ruoli = tk.Frame(self.root, bd=2, relief="groove")
        self.frame_ruoli.grid(row=0, column=2, sticky="nsew", padx=5, pady=5)

        # Configurazione dell'espansione all'interno dei frame
        for frame in [self.frame_personale, self.frame_stanze, self.frame_ruoli]:
            frame.grid_rowconfigure(0, weight=1)
            frame.grid_columnconfigure(0, weight=1)

        # Creazione dell'interfaccia per ciascuna sezione
        self.crea_interfaccia_personale()
        self.crea_interfaccia_stanze()
        self.crea_interfaccia_ruoli()
        # Aggiornamento dei widget con i dati del database
        self.aggiorna_lista_personale()
        self.aggiorna_lista_stanze()
        self.aggiorna_lista_ruoli()

    def crea_interfaccia_personale(self):
        # Configura i widget per la sezione Gestione Personale
        tk.Label(self.frame_personale, text="Gestione Personale", font=("Helvetica", 16)).pack()

        self.label_nome = tk.Label(self.frame_personale, text="Nome:")
        self.label_nome.pack()
        self.entry_nome = tk.Entry(self.frame_personale)
        self.entry_nome.pack()

        self.label_ore = tk.Label(self.frame_personale, text="Ore Settimanali:")
        self.label_ore.pack()
        self.entry_ore = tk.Entry(self.frame_personale)
        self.entry_ore.pack()

        self.label_ruolo_personale = tk.Label(self.frame_personale, text="Ruolo:")
        self.label_ruolo_personale.pack()
        self.listbox_ruoli_personale = tk.Listbox(self.frame_personale, selectmode=tk.MULTIPLE)
        self.listbox_ruoli_personale.pack()

        self.button_aggiungi_personale = tk.Button(self.frame_personale, text="Aggiungi Personale", command=self.aggiungi_personale)
        self.button_aggiungi_personale.pack()

        self.listbox_personale = tk.Listbox(self.frame_personale, width=40)
        self.listbox_personale.pack()

        self.button_rimuovi_personale = tk.Button(self.frame_personale, text="Rimuovi Personale", command=self.rimuovi_personale)
        self.button_rimuovi_personale.pack()

        self.button_modifica_personale = tk.Button(self.frame_personale, text="Modifica Personale", command=self.modifica_personale)
        self.button_modifica_personale.pack()

    def crea_interfaccia_stanze(self):
        # Configura i widget per la sezione Gestione Stanze
        tk.Label(self.frame_stanze, text="Gestione Stanze", font=("Helvetica", 16)).pack()

        self.label_nome_stanza = tk.Label(self.frame_stanze, text="Nome Stanza:")
        self.label_nome_stanza.pack()
        self.entry_nome_stanza = tk.Entry(self.frame_stanze)
        self.entry_nome_stanza.pack()

        self.label_num_persone_ruolo = tk.Label(self.frame_stanze, text="Numero di persone per ruolo:")
        self.label_num_persone_ruolo.pack()
        self.listbox_ruoli_stanze = tk.Listbox(self.frame_stanze, selectmode=tk.SINGLE)
        self.listbox_ruoli_stanze.pack()

        self.entry_num_persone_ruolo = tk.Entry(self.frame_stanze)
        self.entry_num_persone_ruolo.pack()

        self.button_aggiungi_stanza = tk.Button(self.frame_stanze, text="Aggiungi Stanza", command=self.aggiungi_stanza)
        self.button_aggiungi_stanza.pack()

        self.listbox_stanze = tk.Listbox(self.frame_stanze, width=40)
        self.listbox_stanze.pack()

        self.button_rimuovi_stanza = tk.Button(self.frame_stanze, text="Rimuovi Stanza", command=self.rimuovi_stanza)
        self.button_rimuovi_stanza.pack()

    def crea_interfaccia_ruoli(self):
        # Configura i widget per la sezione Gestione Ruoli
        tk.Label(self.frame_ruoli, text="Gestione Ruoli", font=("Helvetica", 16)).pack()

        self.label_nome_ruolo = tk.Label(self.frame_ruoli, text="Nome Ruolo:")
        self.label_nome_ruolo.pack()
        self.entry_nome_ruolo = tk.Entry(self.frame_ruoli)
        self.entry_nome_ruolo.pack()

        self.button_aggiungi_ruolo = tk.Button(self.frame_ruoli, text="Aggiungi Ruolo", command=self.aggiungi_ruolo)
        self.button_aggiungi_ruolo.pack()

        self.listbox_ruoli = tk.Listbox(self.frame_ruoli, width=40)
        self.listbox_ruoli.pack()

        self.button_rimuovi_ruolo = tk.Button(self.frame_ruoli, text="Rimuovi Ruolo", command=self.rimuovi_ruolo)
        self.button_rimuovi_ruolo.pack()

        # Esportazione in Excel
        self.button_export_excel = tk.Button(self.frame_ruoli, text="Esporta Pianificazione in Excel",
                                             command=self.esporta_pianificazione_excel)
        self.button_export_excel.pack()

    def crea_tabelle(self):
        cursor = self.connection.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS dipendenti (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT UNIQUE NOT NULL,
                ore_disponibili INTEGER NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS ruoli (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT UNIQUE NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS dipendenti_ruoli (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                dipendente_id INTEGER NOT NULL,
                ruolo_id INTEGER NOT NULL,
                FOREIGN KEY (dipendente_id) REFERENCES dipendenti(id),
                FOREIGN KEY (ruolo_id) REFERENCES ruoli(id)
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS stanze (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT UNIQUE NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS stanze_ruoli (
                stanza_id INTEGER NOT NULL,
                ruolo_id INTEGER NOT NULL,
                numero_persone INTEGER NOT NULL,
                FOREIGN KEY (stanza_id) REFERENCES stanze(id),
                FOREIGN KEY (ruolo_id) REFERENCES ruoli(id)
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS orari_settimanali (
                giorno TEXT PRIMARY KEY,
                ore_lavorative INTEGER NOT NULL
            )
        """)
        self.connection.commit()

    def aggiungi_personale(self):
        nome = self.entry_nome.get()
        try:
            ore_disponibili = int(self.entry_ore.get())
            ruoli_selezionati = [self.listbox_ruoli_personale.get(i) for i in self.listbox_ruoli_personale.curselection()]
            if nome and ore_disponibili >= 0:
                self.gestione.aggiungi_dipendente(nome, ore_disponibili, ruoli_selezionati)
                self.aggiorna_lista_personale()
                self.entry_nome.delete(0, tk.END)
                self.entry_ore.delete(0, tk.END)
            else:
                messagebox.showerror("Errore", "Nome o ore non validi!")
        except ValueError:
            messagebox.showerror("Errore", "Le ore devono essere un numero intero!")

    def rimuovi_personale(self):
        try:
            nome = self.listbox_personale.get(tk.ACTIVE).split(" - ")[0]
            self.gestione.rimuovi_dipendente(nome)
            self.aggiorna_lista_personale()
        except IndexError:
            messagebox.showerror("Errore", "Seleziona un dipendente dalla lista.")

    def modifica_personale(self):
        try:
            nome_selezionato = self.listbox_personale.get(tk.ACTIVE).split(" - ")[0]
            nuovo_nome = self.entry_nome.get()
            nuove_ore = int(self.entry_ore.get())
            ruoli_selezionati = [self.listbox_ruoli_personale.get(i) for i in self.listbox_ruoli_personale.curselection()]

            if nuovo_nome and nuove_ore >= 0:
                self.gestione.modifica_dipendente(nome_selezionato, nuovo_nome, nuove_ore, ruoli_selezionati)
                self.aggiorna_lista_personale()
                self.entry_nome.delete(0, tk.END)
                self.entry_ore.delete(0, tk.END)
            else:
                messagebox.showerror("Errore", "Nome o ore non validi!")
        except (ValueError, IndexError):
            messagebox.showerror("Errore", "Seleziona un dipendente e inserisci valori validi.")

    def aggiorna_lista_personale(self):
        self.listbox_personale.delete(0, tk.END)
        for dipendente in self.gestione.dipendenti:
            self.listbox_personale.insert(tk.END, str(dipendente))

    def aggiungi_stanza(self):
        nome_stanza = self.entry_nome_stanza.get()
        try:
            ruolo = self.listbox_ruoli_stanze.get(tk.ACTIVE)
            num_persone = int(self.entry_num_persone_ruolo.get())
            persone_ruoli = {ruolo: num_persone}
            if nome_stanza and num_persone > 0:
                self.gestione.aggiungi_stanza(nome_stanza, persone_ruoli)
                self.aggiorna_lista_stanze()
                self.entry_nome_stanza.delete(0, tk.END)
                self.entry_num_persone_ruolo.delete(0, tk.END)
            else:
                messagebox.showerror("Errore", "Nome della stanza o numero di persone non validi!")
        except ValueError:
            messagebox.showerror("Errore", "Il numero di persone deve essere un numero intero positivo!")

    def rimuovi_stanza(self):
        nome_stanza = self.entry_nome_stanza.get(tk.ACTIVE).split(" - ")[0]
        self.gestione.rimuovi_stanza(nome_stanza)
        self.aggiorna_lista_stanze()
        self.entry_nome_stanza.delete(0, tk.END)

    def aggiorna_lista_stanze(self):
        self.listbox_stanze.delete(0, tk.END)
        for stanza in self.gestione.stanze:
            persone_ruoli = self.gestione.num_persone_stanze_ruolo[stanza]
            descrizione_stanza = f"{stanza}: {', '.join([f'{ruolo} ({num})' for ruolo, num in persone_ruoli.items()])}"
            self.listbox_stanze.insert(tk.END, descrizione_stanza)

    def aggiungi_ruolo(self):
        nome_ruolo = self.entry_nome_ruolo.get()
        if nome_ruolo:
            self.gestione.aggiungi_ruolo(nome_ruolo)
            self.aggiorna_lista_ruoli()
            self.entry_nome_ruolo.delete(0, tk.END)
        else:
            messagebox.showerror("Errore", "Inserisci un nome per il ruolo!")

    def rimuovi_ruolo(self):
        nome_ruolo = self.listbox_ruoli.get(tk.ACTIVE)
        if nome_ruolo:
            self.gestione.rimuovi_ruolo(nome_ruolo)
            self.aggiorna_lista_ruoli()
        else:
            messagebox.showerror("Errore", "Seleziona un ruolo dalla lista!")

    def aggiorna_lista_ruoli(self):
        self.listbox_ruoli.delete(0, tk.END)
        self.listbox_ruoli_personale.delete(0, tk.END)
        self.listbox_ruoli_stanze.delete(0, tk.END)
        for ruolo in self.gestione.ruoli:
            self.listbox_ruoli.insert(tk.END, ruolo.nome)
            self.listbox_ruoli_personale.insert(tk.END, ruolo.nome)
            self.listbox_ruoli_stanze.insert(tk.END, ruolo.nome)

    def esporta_pianificazione_excel(self):
        self.gestione.esporta_pianificazione_excel()
        messagebox.showinfo("Esportazione completata", "La pianificazione settimanale è stata esportata in Excel.")

if __name__ == "__main__":
    root = tk.Tk()
    app = GestioneRistoranteApp(root)
    root.mainloop()
