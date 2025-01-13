import tkinter as tk
from tkinter import filedialog, messagebox
import services.EmployeeService as EMPSERV
import services.RoomService as ROOMSERV
import services.RoleService as ROLESERV
import openpyxl
import os
from datetime import datetime

WEEK_DAYS = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
save_path = ""

def browse_folder(entry_export):
    folder = filedialog.askdirectory()
    if folder:
        global save_path
        save_path = folder
        entry_export.delete(0, tk.END)
        entry_export.insert(0, folder)

def build_schedule(self):
    if not save_path:
        messagebox.showerror("Errore", "Seleziona una cartella di destinazione per l'esportazione")
        return

    # build a dictionary with all the rooms and their needed roles from repo,
    # roles with more needed employees are repeated as many times as needed.
    rooms_roles = {}
    rooms = ROOMSERV.get_rooms()
    for r in rooms:
        rooms_roles[r] = ROOMSERV.explode_roles(ROOMSERV.get_room_roles(r))

    # build a dictionary with all the employees and their working hours from repo
    # the remainig hours will be updated after each assignation
    employees = EMPSERV.get_employees()

    # build a dictionary with all the roles and the available employees from repo
    # each role has a list of employee names
    employees_by_role = {}
    for r in ROLESERV.get_roles():
        employees_by_role[r] = EMPSERV.get_employees_by_role(r)
    
    # this dictionary will contain the assigned employees for each day to avoid double assignation
    assigned_employees = {}

    schedule = {}
    # schedule example structure:
    # schedule = {
    #     "Lunedì": {
    #         room : {[employee, role, hours, empty shift] [employee, role, hours, empty shift] [employee, role, hours, empty shift]}}
    #     "Martedì": {etc...}
    #           }

    for day in WEEK_DAYS:
        schedule[day] = {}
        assigned_employees[day] = []
        for room, roles in rooms_roles.items():
            schedule[day][room] = []
            for role in roles:
                shift = ""
                remaining_hours = 8
                for e in employees_by_role[role]:
                    if remaining_hours == 0:
                        break
                    if e in employees:
                        employee_available_hours = employees[e][1]
                        if employee_available_hours > 0 and e not in assigned_employees[day]: # if the employee is not already assigned that day to another room/role
                            assigned_employees[day].append(e)
                        
                            if employee_available_hours >= remaining_hours: # if the employee has enough hours to cover the remaining hours
                                schedule[day][room].append((e, role, remaining_hours, False)) # assign the employee for the remaining hours
                                employees[e][1] -= remaining_hours # update the remaining hours of this employee   
                                remaining_hours = 0 # no more hours to assign for this role in this room today
                            else:
                                schedule[day][room].append((e, role, employee_available_hours, False)) # assign the employee for all the hours he has available
                                remaining_hours -= employee_available_hours # update the remaining hours for this role in this room today
                                employees[e][1] = 0 # the employee has no more hours available
                    else:
                        continue  
                if remaining_hours > 0: # if there are still hours to assign for this role in this room today
                    schedule[day][room].append(("Nessun dipendente disponibile", role, remaining_hours, True)) # signal that there are no more employees available for this role in this room today    

    schedule["Report"] = [[],[]]
    for day in WEEK_DAYS:
        for room in schedule[day]:
            for shift in schedule[day][room]:
                if shift[3]:
                    schedule["Report"][0].append((day, room, shift[1], shift[2]))
    for e in employees:
        if employees[e][1] > 0:
            schedule["Report"][1].append((e, employees[e][1]))
    return schedule

def build_excel_spreadsheet(schedule):
    wb = openpyxl.Workbook()
    schedule_ws = wb.active
    schedule_ws.title = "Pianificazione Settimanale"

    for day in WEEK_DAYS:
        c = WEEK_DAYS.index(day) * 4 + 1
        schedule_ws.cell(1, c).value = day
        r = 2
        for room in schedule[day]:
            schedule_ws.cell(r, c).value = room
            for shift in schedule[day][room]:
                schedule_ws.cell(r, c+1).value = shift[1]    
                hours = f"{shift[2]} ora" if shift[2] == 1 else f"{shift[2]} ore"
                schedule_ws.cell(r, c+2).value = f"{shift[0]} ({hours})"
                r += 1

    # Create a new worksheet for the report
    report_ws = wb.create_sheet(title="Report")
    report_ws.cell(1, 1).value = "Turni non coperti:"
    r = 2
    p_day = ""
    p_room = ""
    p_role = ""
    for day, room, role, hours in schedule["Report"][0]:
        if day != p_day:
            report_ws.cell(r,1).value = day
            p_day = day
        
        if room != p_room or day != p_day:
            report_ws.cell(r,2).value = room
            p_room = room

        if role != p_role or room != p_room or day != p_day:
            report_ws.cell(r,3).value = role
            p_role = role

        if hours == 1:
            s = "1 ora di questo turno non è coperta"        
        else:
            s = f"{hours} ore di questo turno non sono coperte"

        report_ws.cell(r, 4).value = s
        r += 1
    
    report_ws.cell(r+1, 1).value = "Dipendenti con ore in esubero:"
    r += 2
    for employee in schedule["Report"][1]:
        hours = f"{employee[1]} ora disponibile" if employee[1] == 1 else f"{employee[1]} ore disponibili"
        roles = ", ".join([role for role in EMPSERV.get_roles_by_employee(employee[0])])
        report_ws.cell(r, 1).value = f"{employee[0]} ({roles}) ha ancora {hours}"
        r += 1
   
    # Automatically set the column width based on its content
    for column in schedule_ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        schedule_ws.column_dimensions[column[0].column_letter].width = adjusted_width

    for column in report_ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        if column[0].column_letter != 'A':
            report_ws.column_dimensions[column[0].column_letter].width = adjusted_width    

    # Generate filename with current date and time
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Pianificazione_Settimanale_{current_time}.xlsx"

    file_path = os.path.join(save_path, filename)
    wb.save(file_path)

def export_schedule(self):
    schedule = build_schedule(self)
    build_excel_spreadsheet(schedule)
    messagebox.showinfo("Esportazione completata", "La pianificazione settimanale è stata esportata in Excel")